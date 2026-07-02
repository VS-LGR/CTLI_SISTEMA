#!/usr/bin/env python3
"""Extrai fórmulas e valores da aba Metr. Legal do RE-7.2B."""
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_PATH = os.path.join(
    os.path.expanduser("~"),
    "Downloads",
    "RE-7.2B Certificado de Calibração Rev. 00 Matriz (3).xlsm",
)


def cell_info(ws_f, ws_v, row, col):
    f = ws_f.cell(row, col).value
    v = ws_v.cell(row, col).value
    return {
        "col": col,
        "value": v,
        "formula": f if isinstance(f, str) and f.startswith("=") else None,
    }


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    wb_f = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    ws_f = wb_f["Metr. Legal"]
    ws_v = wb_v["Metr. Legal"]

    headers = {}
    for c in range(1, 124):
        h = ws_f.cell(2, c).value
        if h:
            headers[c] = str(h).strip()

    certs = []
    for r in range(3, 20):
        name = ws_v.cell(r, 1).value
        if not name or str(name) in ("0", "None"):
            continue
        certs.append({"row": r, "name": str(name)})

    # Template formulas from first cert row (row 4)
    template_row = certs[0]["row"] if certs else 4
    template = {}
    for c, h in headers.items():
        info = cell_info(ws_f, ws_v, template_row, c)
        if info["formula"] or info["value"] not in (None, "", 0):
            template[h] = info

    # P1 block column mapping (cols 19-27 approx)
    p1_cols = {headers.get(c, f"col{c}"): c for c in range(19, 28) if c in headers}

    out = {
        "source": os.path.basename(path),
        "headers": headers,
        "certificates": certs,
        "p1_column_map": p1_cols,
        "template_row": template_row,
        "template_sample": {k: v for k, v in list(template.items())[:80]},
        "class_formulas": {
            headers.get(c, f"col{c}"): cell_info(ws_f, ws_v, template_row, c)
            for c in range(9, 18) if c in headers
        },
        "resultado_final": cell_info(ws_f, ws_v, template_row, 123),
        "excentricidade": {
            headers.get(c, f"col{c}"): cell_info(ws_f, ws_v, template_row, c)
            for c in range(108, 123) if c in headers
        },
    }

    # Per-cert P1 results for validation
    for cert in certs[:6]:
        r = cert["row"]
        cert["data"] = {
            "divisao_verificacao": ws_v.cell(r, 3).value,
            "faixa_indicacao": ws_v.cell(r, 6).value,
            "resultado_p1": ws_v.cell(r, 27).value,
            "p1_erro": ws_v.cell(r, 26).value,
            "ajuste_media_p1": ws_v.cell(r, 25).value,
            "cliente_erro_p1": ws_v.cell(r, 21).value,
            "cliente_eu_p1": ws_v.cell(r, 22).value,
            "tol_pos_p1": ws_v.cell(r, 23).value,
            "tol_neg_p1": ws_v.cell(r, 24).value,
            "portaria236_p1": ws_v.cell(r, 20).value,
            "resultado_final": ws_v.cell(r, 123).value,
            "resultado_parcial": ws_v.cell(r, 107).value,
            "formulas_p1": {
                "e_p1": ws_f.cell(r, 19).value,
                "portaria": ws_f.cell(r, 20).value,
                "cliente_erro": ws_f.cell(r, 21).value,
                "cliente_eu": ws_f.cell(r, 22).value,
                "tol_pos": ws_f.cell(r, 23).value,
                "tol_neg": ws_f.cell(r, 24).value,
                "ajuste": ws_f.cell(r, 25).value,
                "erro": ws_f.cell(r, 26).value,
                "resultado": ws_f.cell(r, 27).value,
            },
            "formulas_classe": {
                "classe_iii": ws_f.cell(r, 9).value,
                "classe_ii": ws_f.cell(r, 12).value,
                "classe_i": ws_f.cell(r, 15).value,
                "parcial": ws_f.cell(r, 17).value,
                "resultado_classe": ws_f.cell(r, 18).value,
            },
            "formulas_final": ws_f.cell(r, 123).value,
        }

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "src",
        "lib",
        "certificateCalculations",
        "__fixtures__",
        "metr-legal-extract.json",
    )
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)
    print(f"Written {out_path}")
    print(f"Certs: {[c['name'] for c in certs]}")
    print("P1 formulas:", certs[0]["data"]["formulas_p1"] if certs else "none")


if __name__ == "__main__":
    main()
