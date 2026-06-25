#!/usr/bin/env python3
"""
Extrai valores de referência da RE-7.2B Matriz (2).xlsm para auditoria.
Uso: python scripts/extract-xlsm-golden.py [caminho.xlsm]

Nota: Certificado-RBC reflete só o certificado ativo na planilha; display é
recalculado aqui com MROUND (mesma regra de certificateDisplayRounding.js).
"""
import json
import math
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

UE_DISPLAY_FACTOR = 4.4


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    s = str(v).replace(",", ".").strip()
    if s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def veff_display(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip() in ("∞", "inf", "Infinito"):
        return "∞"
    n = num(v)
    if n is None:
        return None
    if n > 99 or n == float("inf"):
        return "∞"
    return str(int(n))


def mround(value, multiple):
    v = num(value)
    m = num(multiple)
    if v is None or m is None or m <= 0:
        return None
    if v == 0:
        return 0.0
    steps = round(v / m + 1e-12)
    return steps * m


def display_expanded_uncertainty(ue_calc, resolution):
    ue = num(ue_calc)
    d = num(resolution)
    if ue is None or d is None or d <= 0:
        return ue
    base = max(ue, d)
    adjusted = base + (d / 10) * UE_DISPLAY_FACTOR
    return mround(adjusted, d)


def display_indication_error(average, reference, resolution):
    avg = num(average)
    ref = num(reference)
    d = num(resolution)
    if avg is None or ref is None or d is None:
        return None
    return mround(avg, d) - mround(ref, d)


def find_xlsm(path_arg):
    if path_arg and os.path.isfile(path_arg):
        return path_arg
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    for f in os.listdir(downloads):
        if "RE-7.2B" in f and "Matriz (2)" in f and f.endswith(".xlsm"):
            return os.path.join(downloads, f)
    return None


def px_columns(point_number):
    """Colunas pós-ajuste nas abas P1..P10 (data_only)."""
    if point_number == 1:
        return {
            "reference": 8,
            "average": 28,
            "ua": 29,
            "up": 30,
            "ud": 31,
            "ue": 32,
            "ur": 33,
            "veff": 34,
            "k": 35,
            "uc": 36,
            "expanded": 37,
            "load_batch": 40,
            "up_lc": None,
        }
    return {
        "reference": 8,
        "average": 27,
        "ua": 28,
        "up": 29,
        "ud": 30,
        "ue": 31,
        "ur": 32,
        "veff": 33,
        "k": 34,
        "uc": 35,
        "expanded": 36,
        "load_batch": 40,
        "up_lc": 35 if point_number == 2 else 34,
    }


def extract(path):
    wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    pre = wb_v["PREENCHER"]
    cert_sheet = wb_v["Certificado-RBC"]

    certs = []
    for r in range(5, 15):
        name = pre.cell(r, 1).value
        if not name or str(name) in ("0", "Informações"):
            continue
        n_points = pre.cell(r, 5).value
        if n_points is None:
            continue
        certs.append({"name": str(name), "n_points": int(n_points), "pre_row": r})

    output = {"source": os.path.basename(path), "certificates": []}

    for cert in certs:
        entry = {
            "id": cert["name"].replace("/", "-").replace(" ", "-").lower(),
            "name": cert["name"],
            "n_points": cert["n_points"],
            "points": [],
        }

        for pn in range(1, min(cert["n_points"], 10) + 1):
            sheet = f"P{pn}"
            if sheet not in wb_v.sheetnames:
                continue
            ws = wb_v[sheet]
            data_row = None
            for r in range(7, 25):
                if ws.cell(r, 1).value == cert["name"]:
                    data_row = r
                    break
            if data_row is None:
                continue

            cols = px_columns(pn)
            resolution = num(cert_sheet.cell(108 + pn, 35).value)

            ref = num(ws.cell(data_row, cols["reference"]).value)
            avg = num(ws.cell(data_row, cols["average"]).value)
            uc = num(ws.cell(data_row, cols["uc"]).value)
            ue = num(ws.cell(data_row, cols["expanded"]).value)
            k = num(ws.cell(data_row, cols["k"]).value)
            veff_raw = num(ws.cell(data_row, cols["veff"]).value)

            lote_val = ws.cell(data_row, cols["load_batch"]).value
            use_load_batch = str(lote_val or "").lower() in ("sim", "s", "yes")

            calc = {
                "point_number": pn,
                "reference": ref,
                "average": avg,
                "ua": num(ws.cell(data_row, cols["ua"]).value),
                "up": num(ws.cell(data_row, cols["up"]).value),
                "ud": num(ws.cell(data_row, cols["ud"]).value),
                "ue": num(ws.cell(data_row, cols["ue"]).value),
                "ur": num(ws.cell(data_row, cols["ur"]).value),
                "up_lc": num(ws.cell(data_row, cols["up_lc"]).value) if cols["up_lc"] and pn >= 2 else 0,
                "combined_uncertainty": uc,
                "expanded_uncertainty": ue,
                "coverage_factor": k,
                "veff_raw": veff_raw,
                "use_load_batch": use_load_batch,
                "resolution": resolution,
            }

            display = {
                "reference": mround(ref, resolution) if ref is not None else None,
                "average": mround(avg, resolution) if avg is not None else None,
                "indication_error": display_indication_error(avg, ref, resolution),
                "expanded_uncertainty": display_expanded_uncertainty(ue, resolution),
                "veff": veff_display(veff_raw),
                "k": k,
                "resolution": resolution,
            }

            entry["points"].append({"calc": calc, "display": display})

        if entry["points"]:
            output["certificates"].append(entry)

    return output


def main():
    path = find_xlsm(sys.argv[1] if len(sys.argv) > 1 else None)
    if not path:
        print("Arquivo RE-7.2B Matriz (2).xlsm não encontrado.", file=sys.stderr)
        sys.exit(1)
    data = extract(path)
    out = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "src",
        "lib",
        "certificateCalculations",
        "__fixtures__",
        "xlsm-matriz2-audit.json",
    )
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Escrito: {out} ({len(data['certificates'])} certificados)")


if __name__ == "__main__":
    main()
